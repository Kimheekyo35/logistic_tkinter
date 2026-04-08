import fitz
from pathlib import Path
import re
from openpyxl.styles import Alignment
import openpyxl as oxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Color


def _build_merged_pdf_pattern(country: str) -> re.Pattern:
    # Support merged PDFs produced by different brand crawlers.
    prefixes = [
        r"NUMBUZIN_\d{4}_\d{2}_\d{2}",
        r"TAGE_\d{4}_\d{2}_\d{2}",
        r"TAGE_TEST_\d{4}_\d{2}_\d{2}",
        r"TAGE_FM_\d{4}_\d{2}_\d{2}",
        r"TAGE_FM_TEST_\d{4}_\d{2}_\d{2}",
    ]
    prefix_group = "|".join(prefixes)
    return re.compile(
        rf"^(?:{prefix_group})_{re.escape(country)}_merged_\d+\.pdf$",
        re.IGNORECASE,
    )

def pdf_to_text(folder_path: str, country: str, output_path: str):
    wb = oxl.Workbook()
    ws = wb.active

    a1 = ws["A1"]
    a1.value = "Order ID"
    a1.fill = PatternFill(fill_type='solid',
                          fgColor=Color('789ABC'))

    order_id_list = []
    folder = Path(folder_path)
    pattern = _build_merged_pdf_pattern(country)
    files = sorted(
        file for file in folder.iterdir()
        if file.is_file() and pattern.match(file.name)
    )

    for file in files:
        doc = fitz.open(file)
        for row, page_num in enumerate(range(len(doc)), start=1):
            page = doc.load_page(page_num)
            text = page.get_text()
            for line in text.splitlines():
                order_id = line.replace("Order ID: ", "")
                order_id_list.append(order_id)
                break

    for row, order_id in enumerate(order_id_list, start=1):
        ws["A" + str(row + 1)] = order_id

    wb.save(output_path)
    return output_path
